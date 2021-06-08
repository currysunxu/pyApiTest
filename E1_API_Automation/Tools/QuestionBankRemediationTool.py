import json
import uuid
from enum import Enum

import jmespath
import json_tools
import openpyxl
import requests
from hamcrest import assert_that

from E1_API_Automation.Business.NGPlatform.ContentMapQueryEntity import ContentMapQueryEntity
from E1_API_Automation.Business.NGPlatform.ContentMapService import ContentMapService
from E1_API_Automation.Business.NGPlatform.ContentRepoService import ContentRepoService
from E1_API_Automation.Business.PipelinePublish.AEMService import AEMService
from E1_API_Automation.Business.PipelinePublish.PipelinePublishUtils.PipelinePublishConstants import \
    PipelinePublishConstants
from E1_API_Automation.Test_Data.PipelinePublishData import AEMData


class RemediationEntity:
    def __init__(self, test_type, skill_set):
        self.__test_type = test_type
        self.__skill_set = skill_set
        self.__activity_key = None
        self.__activity_title = None
        self.__missed_question_key = None
        self.__missed_question_tags = None
        self.__recommend_questions = None

    @property
    def test_type(self):
        return self.__test_type

    @test_type.setter
    def test_type(self, test_type):
        self.__test_type = test_type

    @property
    def skill_set(self):
        return self.__skill_set

    @skill_set.setter
    def skill_set(self, skill_set):
        self.__skill_set = skill_set

    @property
    def activity_key(self):
        return self.__activity_key

    @activity_key.setter
    def activity_key(self, activity_key):
        self.__activity_key = activity_key

    @property
    def activity_title(self):
        return self.__activity_title

    @activity_title.setter
    def activity_title(self, activity_title):
        self.__activity_title = activity_title

    @property
    def missed_question_key(self):
        return self.__missed_question_key

    @missed_question_key.setter
    def missed_question_key(self, missed_question_key):
        self.__missed_question_key = missed_question_key

    @property
    def missed_question_tags(self):
        return self.__missed_question_tags

    @missed_question_tags.setter
    def missed_question_tags(self, missed_question_tags):
        self.__missed_question_tags = missed_question_tags

    @property
    def recommend_questions(self):
        return self.__recommend_questions

    @recommend_questions.setter
    def recommend_questions(self, recommend_questions):
        self.__recommend_questions = recommend_questions


class RemediationQueryType(Enum):
    TypeScoring = 'SCORING'
    TypeExactMatch = 'EXACT_MATCH'
    TypePrimarySkillBased = 'PRIMARY_SKILL_BASED'


# get midterm/final term questions' remediation questions
def get_mistaken_question_remediations(source_aem_env, release_program, check_book_list, region_ach):
    for check_book in check_book_list:
        print('-------------------------Start check book: {0}'.format(check_book))
        source_course = AEMData.CourseData[release_program]['source-name']
        book_content_path = '{0}/{1}/{2}'.format(source_course, region_ach, check_book)

        content_map_service = ContentMapService()
        content_map_course = AEMData.CourseData[release_program]['target-name']
        content_map_entity = ContentMapQueryEntity(content_map_course, region_ach=region_ach)
        content_map_tree_response = content_map_service.post_content_map_query_tree(content_map_entity)
        assert_that(content_map_tree_response.status_code == 200)
        content_map_book_tree = jmespath.search('children[? contentPath == \'{0}\']|[0]'.format(book_content_path),
                                                content_map_tree_response.json())
        content_revision = content_map_book_tree['contentRevision']

        book_question_remediation_list = []
        aem_service = AEMService(AEMData.AEMHost[source_aem_env])
        aem_book_response = aem_service.get_aem_book(source_course, check_book)
        aem_units_dict = aem_book_response.json()['units']
        for unit_key in aem_units_dict.keys():
            aem_unit_value_dict = aem_units_dict[unit_key]
            aem_unit_path = aem_unit_value_dict['path']
            aem_unit_term_dict = None
            if 'midTerm' in aem_unit_value_dict.keys():
                test_type = 'midTerm'
                aem_unit_term_dict = aem_unit_value_dict['midTerm']
            elif 'final' in aem_unit_value_dict.keys():
                test_type = 'finalTerm'
                aem_unit_term_dict = aem_unit_value_dict['final']

            if aem_unit_term_dict is not None:
                for skill_set in aem_unit_term_dict.keys():
                    aem_unit_term_skill_list = aem_unit_term_dict[skill_set]

                    for aem_unit_term_activity in aem_unit_term_skill_list:
                        aem_unit_term_activity_url = aem_unit_term_activity[PipelinePublishConstants.FIELD_URL]
                        aem_unit_term_activity_detail_response = requests.get(aem_unit_term_activity_url)
                        if aem_unit_term_activity_detail_response.status_code != 200:
                            print("get aem_unit_term_activity_url failed: " + aem_unit_term_activity_url)
                        assert_that(aem_unit_term_activity_detail_response.status_code == 200)

                        # version 1: implement the es query by myself
                        # book_question_remediation_list.extend(get_activity_question_remediation_list(
                        #     aem_unit_term_activity_detail_response.json(),
                        #     region_ach,
                        #     aem_unit_path,
                        #     test_type,
                        #     skill_set,
                        #     content_revision))
                        book_question_remediation_list.extend(
                            get_activity_question_remediation_list_scoring_exact_match(
                                aem_unit_term_activity_detail_response.json(),
                                region_ach,
                                aem_unit_path,
                                test_type,
                                skill_set,
                                content_revision))

        # version 1: implement the es query by myself
        # save_to_excel(check_book, book_question_remediation_list)
        save_to_excel_with_both_scoring_exact_match(check_book, book_question_remediation_list)


# get remediation by scoring, the feature is implemented by myself as there's no API provided at first
def get_activity_question_remediation_list(aem_term_activity, region, unit_path, test_type, skill_set,
                                           content_revision):
    activity_remediation_list = []
    activity_title = aem_term_activity['Title']
    aem_activity_question_list = aem_term_activity['Questions']
    is_listening_reading = False
    if skill_set in ('listening', 'reading'):
        is_listening_reading = True

    if not is_listening_reading:
        for aem_activity_question in aem_activity_question_list:
            es_query_result_list = query_es_by_question_tags(aem_activity_question, region, unit_path,
                                                             activity_title, content_revision,
                                                             is_listening_reading)
            remediationEntity = RemediationEntity(test_type, skill_set)
            remediationEntity.activity_key = aem_term_activity['Key']
            remediationEntity.missed_question_key = aem_activity_question['key']
            aem_question_tags = aem_activity_question['body']['tags']
            if len(aem_question_tags) > 0:
                remediationEntity.missed_question_tags = remove_non_used_tag_fields(aem_question_tags[0])
            remediationEntity.recommend_questions = es_query_result_list
            activity_remediation_list.append(remediationEntity)
    else:
        aem_reading_listing_question = aem_activity_question_list[0]
        es_query_result_list = query_es_by_question_tags(aem_reading_listing_question, region, unit_path,
                                                         activity_title,
                                                         content_revision, is_listening_reading)
        remediationEntity = RemediationEntity(test_type, skill_set)
        remediationEntity.activity_key = aem_term_activity['Key']
        remediationEntity.missed_question_key = aem_term_activity['Key']
        aem_question_tags = aem_reading_listing_question['body']['tags']
        if len(aem_question_tags) > 0:
            remediationEntity.missed_question_tags = remove_non_used_tag_fields(aem_question_tags[0])
        remediationEntity.recommend_questions = es_query_result_list
        activity_remediation_list.append(remediationEntity)
    return activity_remediation_list


def query_es_by_question_tags(aem_activity_question, region, unit_path, activity_title, content_revision,
                              is_listening_reading):
    es_query_result_list = []
    question_query_es_body = construct_question_tags_query_body(aem_activity_question, region,
                                                                unit_path, activity_title,
                                                                content_revision)
    question_query_es_body = json.dumps(question_query_es_body)
    print(question_query_es_body)
    es_search_response = requests.post('http://internal-elasticsearch-stg.english1.cn:9200/kt-remediation-stg/_search',
                                       question_query_es_body,
                                       headers={"Content-Type": "application/json"})
    es_hits = es_search_response.json()['hits']['hits']
    if not is_listening_reading:
        # for grammar and vocab, get 3 recommend questions for each missed question
        for i in range(3):
            es_query_result_list.append(es_hits[i]['_source'])
    else:
        # for listening and reading, get 1 recommend activity for missed question
        es_query_result_list.append(es_hits[0]['_source'])
    return es_query_result_list


# get midterm/final term remediation list using provided API
def get_activity_question_remediation_list_scoring_exact_match(aem_term_activity, region, unit_path, test_type,
                                                               skill_set, content_revision):
    activity_remediation_list = []
    activity_title = aem_term_activity['Title']
    aem_activity_question_list = aem_term_activity['Questions']
    is_listening_reading = False
    if skill_set in ('listening', 'reading'):
        is_listening_reading = True

    if not is_listening_reading:
        for aem_activity_question in aem_activity_question_list:
            es_recommend_scoring_exact_match_dict \
                = query_scoring_exact_match_api_by_question_tags(aem_activity_question, region, unit_path,
                                                                 activity_title, content_revision,
                                                                 is_listening_reading)
            remediationEntity = RemediationEntity(test_type, skill_set)
            remediationEntity.activity_key = aem_term_activity['Key']
            remediationEntity.activity_title = activity_title
            remediationEntity.missed_question_key = aem_activity_question['key']
            aem_question_tags = aem_activity_question['body']['tags']
            if len(aem_question_tags) > 0:
                remediationEntity.missed_question_tags = remove_non_used_tag_fields(aem_question_tags[0])
            remediationEntity.recommend_questions = es_recommend_scoring_exact_match_dict
            activity_remediation_list.append(remediationEntity)
    else:
        aem_reading_listing_question = aem_activity_question_list[0]
        es_recommend_scoring_exact_match_dict \
            = query_scoring_exact_match_api_by_question_tags(aem_reading_listing_question, region, unit_path,
                                                             activity_title,
                                                             content_revision, is_listening_reading)
        remediationEntity = RemediationEntity(test_type, skill_set)
        remediationEntity.activity_key = aem_term_activity['Key']
        remediationEntity.activity_title = activity_title
        remediationEntity.missed_question_key = aem_term_activity['Key']
        aem_question_tags = aem_reading_listing_question['body']['tags']
        if len(aem_question_tags) > 0:
            remediationEntity.missed_question_tags = remove_non_used_tag_fields(aem_question_tags[0])
        remediationEntity.recommend_questions = es_recommend_scoring_exact_match_dict
        activity_remediation_list.append(remediationEntity)
    return activity_remediation_list


def query_scoring_exact_match_api_by_question_tags(aem_activity_question, region, unit_path, activity_title,
                                                   content_revision, is_listening_reading):
    remediation_query_es_body = construct_remediation_api_query_body(aem_activity_question, region,
                                                                     unit_path, activity_title,
                                                                     content_revision)
    remediation_query_es_body = json.dumps(remediation_query_es_body)
    print(remediation_query_es_body)
    remediation_query_es_url = 'https://internal-ktsvc-stg.english1.cn/remediation/api/private/v1/remediation?strategy={0}'
    if not is_listening_reading:
        # for grammar and vocab, get 3 recommend questions for each missed question
        remediation_query_es_url = remediation_query_es_url + '&size=3'
    else:
        # for listening and reading, get 1 recommend activity for missed question
        remediation_query_es_url = remediation_query_es_url + '&size=1'

    scoring_query_url = remediation_query_es_url.format(RemediationQueryType.TypeScoring.value)
    exact_match_query_url = remediation_query_es_url.format(RemediationQueryType.TypeExactMatch.value)
    primary_skill_based_query_url = remediation_query_es_url.format(RemediationQueryType.TypePrimarySkillBased.value)
    scoring_query_response = requests.get(scoring_query_url,
                                          data=remediation_query_es_body,
                                          headers={"Content-Type": "application/json"})
    exact_match_query_response = requests.get(exact_match_query_url,
                                              data=remediation_query_es_body,
                                              headers={"Content-Type": "application/json"})
    primary_skill_based_query_response = requests.get(primary_skill_based_query_url,
                                                      data=remediation_query_es_body,
                                                      headers={"Content-Type": "application/json"})

    content_repo_service = ContentRepoService()

    scoring_recommend_question_list = append_activity_question_tags(content_repo_service,
                                                                    scoring_query_response.json())
    exact_match_recommend_question_list = append_activity_question_tags(content_repo_service,
                                                                        exact_match_query_response.json())
    primary_skill_based_question_list = append_activity_question_tags(content_repo_service,
                                                                      primary_skill_based_query_response.json())

    es_recommend_scoring_exact_match_dict = {}
    es_recommend_scoring_exact_match_dict[RemediationQueryType.TypeScoring.value] = scoring_recommend_question_list
    es_recommend_scoring_exact_match_dict[
        RemediationQueryType.TypeExactMatch.value] = exact_match_recommend_question_list
    es_recommend_scoring_exact_match_dict[
        RemediationQueryType.TypePrimarySkillBased.value] = primary_skill_based_question_list
    return es_recommend_scoring_exact_match_dict


def append_activity_question_tags(content_repo_service, es_query_result_list):
    activity_detail_list = content_repo_service.get_activities(es_query_result_list).json()

    for es_query_result in es_query_result_list:
        activity_content_id = es_query_result['contentId']
        recommend_question_key = es_query_result['includedQuestions'][0]
        activity_detail_json = jmespath.search('[? contentId == \'{0}\']|[0]'.format(activity_content_id),
                                               activity_detail_list)
        recommend_activity_title = activity_detail_json['data']['Title']
        recommend_question_json = jmespath.search('data.Questions[?key == \'{0}\']|[0]'.format(recommend_question_key),
                                                  activity_detail_json)
        recommend_question_tags = recommend_question_json['body']['tags'][0]
        es_query_result['recommendActivityTitle'] = recommend_activity_title
        es_query_result['includedQuestionTags'] = recommend_question_tags
    return es_query_result_list


def construct_remediation_api_query_body(aem_term_question, region, unit_path, activity_title,
                                         content_revision):
    body_dict = {}
    body_dict['activityKeyRaw'] = str(uuid.uuid1())
    body_dict['activityTitle'] = activity_title
    body_dict['originTags'] = aem_term_question['body']['tags'][0]
    body_dict['activityRevisionRaw'] = content_revision
    slash_index = unit_path.find('/')
    body_dict['activityTagKey'] = unit_path[:slash_index] + '/' + region + unit_path[slash_index:]
    return body_dict


def construct_question_tags_query_body(aem_term_question, region, unit_path, activity_title, content_revision):
    es_query_body_dict = {}
    question_query_dict = {}
    question_tags = aem_term_question['body']['tags'][0]
    cefr = question_tags['cefr'][0]
    primary_skill_set = question_tags['primary-skill-set'][0]

    must_list = []
    must_list.append(construct_term_dict('tags', primary_skill_set, 1.0))
    must_list.append(construct_match_activity_tag(unit_path))
    must_list.append(construct_term_dict('region', region, 1.0))
    must_list.append(construct_term_dict('activityRevisionRaw', content_revision, 1.0))

    filter_list = []
    filter_list.append(construct_term_dict('tags', cefr, 1.0))

    should_list = []
    should_list.append(construct_match_activity_title(activity_title))

    primary_sub_skill_tag_values = get_question_tag_values(question_tags, 'primary-sub-skill-')
    primary_lexical_set_tag_values = get_question_tag_values(question_tags, 'primary-lexical-set-')
    secondary_skill_set_tag_values = get_question_tag_values(question_tags, 'secondary-skill-set')
    secondary_sub_skill_tag_values = get_question_tag_values(question_tags, 'secondary-sub-skill-')
    secondary_lexical_set_tag_values = get_question_tag_values(question_tags, 'secondary-lexical-set-')
    other_tag_values = get_other_question_tag_values(question_tags)

    should_list.extend(construct_should_term(primary_sub_skill_tag_values, 2.0))
    should_list.extend(construct_should_term(primary_lexical_set_tag_values, 0.8))

    if len(secondary_skill_set_tag_values) > 0:
        should_list.extend(construct_should_term(secondary_skill_set_tag_values, 1.0))
        should_list.extend(construct_should_term(secondary_sub_skill_tag_values, 0.8))
    should_list.extend(construct_should_term(secondary_lexical_set_tag_values, 0.6))
    should_list.extend(construct_should_term(other_tag_values, 0.5))

    question_query_dict['must'] = must_list
    question_query_dict['filter'] = filter_list
    question_query_dict['should'] = should_list
    question_query_dict['adjust_pure_negative'] = True
    question_query_dict['boost'] = 1.0

    es_query_body_dict['query'] = {}
    es_query_body_dict['query']['bool'] = question_query_dict
    return es_query_body_dict


def construct_should_term(tag_values, boost):
    should_term_list = []
    for tag_value in tag_values:
        should_term_list.append(construct_term_dict('tags', tag_value, boost))
    return should_term_list


def get_question_tag_values(question_tags, tag_name):
    tag_values = []
    for key in question_tags.keys():
        if key.startswith(tag_name):
            tag_values.extend(question_tags[key])
    return tag_values


def get_other_question_tag_values(question_tags):
    tag_values = []
    for key in question_tags.keys():
        if is_other_tag(key):
            tag_values.extend(question_tags[key])
    return tag_values


def is_other_tag(tag_name):
    key_tag_list = ['cefr', 'primary-skill-set', 'primary-sub-skill-', 'primary-lexical-set-', 'secondary-skill-set',
                    'secondary-sub-skill-', 'secondary-lexical-set-']
    is_other_tag = True
    for key_tag in key_tag_list:
        if tag_name.startswith(key_tag):
            is_other_tag = False
            break
    return is_other_tag


def construct_term_dict(field_name, field_value, boost):
    field_dict = {}

    if field_name == 'region':
        field_value = field_value.upper()
    field_dict['value'] = field_value
    field_dict['boost'] = boost

    term_dict = {}
    term_dict[field_name] = field_dict

    result_dict = {}
    result_dict['term'] = term_dict
    return result_dict


def construct_match_activity_tag(activity_path):
    activity_tag_key_dict = {}
    activity_tag_key_dict['query'] = activity_path
    activity_tag_key_dict['slop'] = 0
    activity_tag_key_dict['max_expansions'] = 50
    activity_tag_key_dict['boost'] = 1.0

    match_phrase_prefix = {}
    match_phrase_prefix['activityTagKey'] = activity_tag_key_dict

    result_dict = {}
    result_dict['match_phrase_prefix'] = match_phrase_prefix
    return result_dict


def construct_match_activity_title(activity_title):
    activity_title_dict = {}
    activity_title_dict['query'] = activity_title
    activity_title_dict['operator'] = 'OR'
    activity_title_dict['prefix_length'] = 0
    activity_title_dict['max_expansions'] = 50
    activity_title_dict['fuzzy_transpositions'] = True
    activity_title_dict['lenient'] = False
    activity_title_dict['zero_terms_query'] = 'NONE'
    activity_title_dict['auto_generate_synonyms_phrase_query'] = True
    activity_title_dict['boost'] = 1.0

    match_dict = {}
    match_dict['activityTitle'] = activity_title_dict

    result_dict = {}
    result_dict['match'] = match_dict
    return result_dict


def remove_non_used_tag_fields(tag_dict):
    if 'intent' in tag_dict.keys():
        tag_dict.pop('intent')
    if 'outcome' in tag_dict.keys():
        tag_dict.pop('outcome')
    # tag_dict.pop('cefr')
    # tag_dict.pop('age-from')
    # tag_dict.pop('age-to')
    if 'rows' in tag_dict.keys():
        tag_dict.pop('rows')
    if 'cols' in tag_dict.keys():
        tag_dict.pop('cols')
    return tag_dict


def save_to_excel(book_name, remediation_list):
    file_name = 'recommendQuestions.xlsx'
    try:
        mywb = openpyxl.load_workbook(file_name)
        mysheet = mywb.create_sheet()
    except:
        mywb = openpyxl.Workbook()
        mysheet = mywb.active

    mysheet.title = book_name
    mysheet.cell(row=1, column=1).value = 'Test Type'
    mysheet.cell(row=1, column=2).value = 'Skill Set'
    mysheet.cell(row=1, column=3).value = 'Activity Key'
    mysheet.cell(row=1, column=4).value = 'Missed Question Key'
    mysheet.cell(row=1, column=5).value = 'Missed Question Tags'
    mysheet.cell(row=1, column=6).value = 'Recommend Question/Activity'
    mysheet.cell(row=1, column=7).value = 'Recommend Question/Activity Tags'
    mysheet.cell(row=1, column=8).value = 'Tag Difference'

    row_number = 2
    for i in range(len(remediation_list)):
        remediation_entity = remediation_list[i]
        mysheet.cell(row=(i + row_number), column=1).value = remediation_entity.test_type
        mysheet.cell(row=(i + row_number), column=2).value = remediation_entity.skill_set
        # mysheet.cell(row=(i + row_number), column=3).value = remediation_entity.activity_key
        # mysheet.cell(row=(i + row_number), column=4).value = remediation_entity.missed_question_key
        recommend_question_size = len(remediation_entity.recommend_questions)
        for j in range(recommend_question_size):
            recommend_question = remediation_entity.recommend_questions[j]
            recommend_question_key = recommend_question['questionKey']
            recommend_question_tags = remove_non_used_tag_fields(recommend_question['originTags'])

            # for listening and reading, recommend activity key
            if remediation_entity.skill_set in ('listening', 'reading'):
                # question key will be like "/content/adam/courses/highflyers/book-2/unit-6/activities/questionbank/maze-maze/grammar-b-q2-2/jcr:content/maze"
                # need to get activity key as "/content/adam/courses/highflyers/book-2/unit-6/activities/questionbank/maze-maze"
                indicator = '/questionbank/'
                indicator_index = recommend_question_key.find(indicator)
                question_bank_end_index = indicator_index + len(indicator)
                left_content = recommend_question_key[question_bank_end_index:]
                activity_name_end_index = left_content.find('/')
                recommend_question_key = recommend_question_key[:question_bank_end_index + activity_name_end_index]
            mysheet.cell(row=(i + row_number + j), column=3).value = remediation_entity.activity_key
            mysheet.cell(row=(i + row_number + j), column=4).value = remediation_entity.missed_question_key
            missed_question_tags_dict = remediation_entity.missed_question_tags
            missed_question_tags_text = json.dumps(missed_question_tags_dict)
            mysheet.cell(row=(i + row_number + j), column=5).value = missed_question_tags_text
            mysheet.cell(row=(i + row_number + j), column=6).value = recommend_question_key
            recommend_question_tags_text = json.dumps(recommend_question_tags)
            mysheet.cell(row=(i + row_number + j), column=7).value = recommend_question_tags_text

            key_diff_list = get_tag_dict_difference(remediation_entity.missed_question_tags,
                                                    recommend_question_tags)

            mysheet.cell(row=(i + row_number + j), column=8).value = "\r\n".join(key_diff_list)

        if recommend_question_size > 1:
            row_number = row_number + (recommend_question_size - 1)
    mywb.save(file_name)


def get_listening_reading_activity_key_by_question_key(question_key):
    # question key will be like "/content/adam/courses/highflyers/book-2/unit-6/activities/questionbank/maze-maze/grammar-b-q2-2/jcr:content/maze"
    # need to get activity key as "/content/adam/courses/highflyers/book-2/unit-6/activities/questionbank/maze-maze"
    indicator = '/questionbank/'
    indicator_index = question_key.find(indicator)
    question_bank_end_index = indicator_index + len(indicator)
    left_content = question_key[question_bank_end_index:]
    activity_name_end_index = left_content.find('/')
    question_key = question_key[:question_bank_end_index + activity_name_end_index]
    return question_key


def save_to_excel_with_both_scoring_exact_match(book_name, remediation_list):
    file_name = 'recommendQuestions_with_three_solutions.xlsx'
    try:
        mywb = openpyxl.load_workbook(file_name)
        mysheet = mywb.create_sheet()
    except:
        mywb = openpyxl.Workbook()
        mysheet = mywb.active

    mysheet.title = book_name
    mysheet.cell(row=1, column=1).value = 'Test Type'
    mysheet.cell(row=1, column=2).value = 'Skill Set'
    mysheet.cell(row=1, column=3).value = 'Activity Key'
    mysheet.cell(row=1, column=4).value = 'Activity Title'
    mysheet.cell(row=1, column=5).value = 'Mistaken Question Key'
    mysheet.cell(row=1, column=6).value = 'Mistaken Question Tags'
    mysheet.cell(row=1, column=7).value = 'Scoring Recommend Activity Title'
    mysheet.cell(row=1, column=8).value = 'Scoring Recommend Question/Activity'
    mysheet.cell(row=1, column=9).value = 'Scoring Recommend Question/Activity Tags'
    mysheet.cell(row=1, column=10).value = 'Scoring Tag Difference'
    mysheet.cell(row=1, column=11).value = 'Exact Match Recommend Activity Title'
    mysheet.cell(row=1, column=12).value = 'Exact Match Recommend Question/Activity'
    mysheet.cell(row=1, column=13).value = 'Exact Match Recommend Question/Activity Tags'
    mysheet.cell(row=1, column=14).value = 'Exact Match Tag Difference'
    mysheet.cell(row=1, column=15).value = 'Primary Skill Based Recommend Activity Title'
    mysheet.cell(row=1, column=16).value = 'Primary Skill Based Recommend Question/Activity'
    mysheet.cell(row=1, column=17).value = 'Primary Skill Based Recommend Question/Activity Tags'
    mysheet.cell(row=1, column=18).value = 'Primary Skill Based Tag Difference'

    row_number = 2
    for i in range(len(remediation_list)):
        remediation_entity = remediation_list[i]
        mysheet.cell(row=(i + row_number), column=1).value = remediation_entity.test_type
        mysheet.cell(row=(i + row_number), column=2).value = remediation_entity.skill_set
        # mysheet.cell(row=(i + row_number), column=3).value = remediation_entity.activity_key
        # mysheet.cell(row=(i + row_number), column=4).value = remediation_entity.missed_question_key
        recommend_question_size = len(remediation_entity.recommend_questions[RemediationQueryType.TypeScoring.value])
        for j in range(recommend_question_size):
            scoring_recommend_question = remediation_entity.recommend_questions[RemediationQueryType.TypeScoring.value][
                j]
            scoring_recommend_activity_title = scoring_recommend_question['recommendActivityTitle']
            scoring_recommend_question_key = scoring_recommend_question['includedQuestions'][0]
            scoring_recommend_question_tags = remove_non_used_tag_fields(
                scoring_recommend_question['includedQuestionTags'])

            exact_match_recommend_question = None
            if j < len(remediation_entity.recommend_questions[RemediationQueryType.TypeExactMatch.value]):
                exact_match_recommend_question = \
                    remediation_entity.recommend_questions[RemediationQueryType.TypeExactMatch.value][j]
                exact_match_recommend_activity_title = exact_match_recommend_question['recommendActivityTitle']
                exact_match_recommend_question_key = exact_match_recommend_question['includedQuestions'][0]
                exact_match_recommend_question_tags = remove_non_used_tag_fields(
                    exact_match_recommend_question['includedQuestionTags'])

            primary_skill_based_recommend_question = None
            if j < len(remediation_entity.recommend_questions[RemediationQueryType.TypePrimarySkillBased.value]):
                primary_skill_based_recommend_question = \
                    remediation_entity.recommend_questions[RemediationQueryType.TypePrimarySkillBased.value][j]
                primary_skill_based_recommend_activity_title = primary_skill_based_recommend_question[
                    'recommendActivityTitle']
                primary_skill_based_recommend_question_key = \
                    primary_skill_based_recommend_question['includedQuestions'][0]
                primary_skill_based_recommend_question_tags = remove_non_used_tag_fields(
                    primary_skill_based_recommend_question['includedQuestionTags'])

            # for listening and reading, recommend activity key
            if remediation_entity.skill_set in ('listening', 'reading'):
                scoring_recommend_question_key = get_listening_reading_activity_key_by_question_key(
                    scoring_recommend_question_key)
                if exact_match_recommend_question is not None:
                    exact_match_recommend_question_key = get_listening_reading_activity_key_by_question_key(
                        exact_match_recommend_question_key)
                if primary_skill_based_recommend_question is not None:
                    primary_skill_based_recommend_question_key = get_listening_reading_activity_key_by_question_key(
                        primary_skill_based_recommend_question_key)
            mysheet.cell(row=(i + row_number + j), column=3).value = remediation_entity.activity_key
            mysheet.cell(row=(i + row_number + j), column=4).value = remediation_entity.activity_title
            mysheet.cell(row=(i + row_number + j), column=5).value = remediation_entity.missed_question_key
            missed_question_tags_dict = remediation_entity.missed_question_tags
            missed_question_tags_text = json.dumps(missed_question_tags_dict)
            mysheet.cell(row=(i + row_number + j), column=6).value = missed_question_tags_text
            mysheet.cell(row=(i + row_number + j), column=7).value = scoring_recommend_activity_title
            mysheet.cell(row=(i + row_number + j), column=8).value = scoring_recommend_question_key
            mysheet.cell(row=(i + row_number + j), column=9).value = json.dumps(scoring_recommend_question_tags)

            scoring_tag_diff_list = get_tag_dict_difference(remediation_entity.missed_question_tags,
                                                            scoring_recommend_question_tags)
            mysheet.cell(row=(i + row_number + j), column=10).value = "\r\n".join(scoring_tag_diff_list)

            if exact_match_recommend_question is not None:
                mysheet.cell(row=(i + row_number + j), column=11).value = exact_match_recommend_activity_title
                mysheet.cell(row=(i + row_number + j), column=12).value = exact_match_recommend_question_key
                mysheet.cell(row=(i + row_number + j), column=13).value = json.dumps(
                    exact_match_recommend_question_tags)
                exact_match_tag_diff_list = get_tag_dict_difference(remediation_entity.missed_question_tags,
                                                                    exact_match_recommend_question_tags)
                mysheet.cell(row=(i + row_number + j), column=14).value = "\r\n".join(exact_match_tag_diff_list)

            if primary_skill_based_recommend_question is not None:
                mysheet.cell(row=(i + row_number + j),
                             column=15).value = primary_skill_based_recommend_activity_title
                mysheet.cell(row=(i + row_number + j), column=16).value = primary_skill_based_recommend_question_key
                mysheet.cell(row=(i + row_number + j), column=17).value = json.dumps(
                    primary_skill_based_recommend_question_tags)
                primary_skill_based_tag_diff_list = get_tag_dict_difference(
                    remediation_entity.missed_question_tags,
                    primary_skill_based_recommend_question_tags)
                mysheet.cell(row=(i + row_number + j), column=18).value = "\r\n".join(
                    primary_skill_based_tag_diff_list)

        if recommend_question_size > 1:
            row_number = row_number + (recommend_question_size - 1)
    mywb.save(file_name)


def get_tag_dict_difference(original_tag_dict, compare_tag_dict):
    diff_list = json_tools.diff(original_tag_dict, compare_tag_dict)

    key_diff_list = []
    if len(diff_list) > 0:
        for key in compare_tag_dict.keys():
            if key in original_tag_dict.keys():
                diff_add = set(compare_tag_dict[key]).difference(set(original_tag_dict[key]))
                diff_remove = set(original_tag_dict[key]).difference(set(compare_tag_dict[key]))
                if len(diff_add) > 0 or len(diff_remove) > 0:
                    diff_str = key + ' field value'
                    if len(diff_add) > 0:
                        diff_str = diff_str + ' added: ' + str(diff_add)
                    if len(diff_remove) > 0:
                        diff_str = diff_str + ' removed: ' + str(diff_remove)
                    # diff_str = key + '\'s value added: {0}, removed: {1}'.format(str(diff_add), str(diff_remove))
                    key_diff_list.append(diff_str)
            else:
                diff_str = key + ' tag key not exist in mistaken question tags, value is:' + str(
                    compare_tag_dict[key])
                key_diff_list.append(diff_str)

        for key in original_tag_dict.keys():
            if key not in compare_tag_dict.keys():
                diff_str = key + ' tag key not exist in recommendation question tags, value is:' + str(
                    original_tag_dict[key])
                key_diff_list.append(diff_str)
    return key_diff_list


# this method can be used to compare tags for mistaken question and recommendation question
def check_fetch_tags_for_excel():
    file_name = 'recommendQuestions_withtagsdiffer_withIntendedAllBooks.xlsx'
    my_wb = openpyxl.load_workbook(file_name)
    sheets = my_wb.get_sheet_names()
    for sheet_name in sheets:
        sheet = my_wb.get_sheet_by_name(sheet_name)

        sheet.cell(row=1, column=12).value = 'Intended Recommendation Tags'
        sheet.cell(row=1, column=13).value = 'Intended Recommendation Tags Difference'

        row_index = 0
        for row in sheet.rows:
            row_index = row_index + 1
            if row_index == 1:
                continue
            intend_recommendation = row[9].value
            skill_type = row[1].value
            if intend_recommendation is not None:
                aem_activity_url = 'https://aem.ef.com/sites.html'
                if intend_recommendation.startswith(aem_activity_url):
                    intend_recommendation = intend_recommendation[len(aem_activity_url):]

                if skill_type not in ('listening', 'reading'):
                    last_slash_index = intend_recommendation.rfind('/')
                    aem_activity_url = 'https://aem.ef.com/apps/adam/activity.standard.json' + intend_recommendation[
                                                                                               :last_slash_index]
                else:
                    aem_activity_url = 'https://aem.ef.com/apps/adam/activity.standard.json' + intend_recommendation
                aem_activity_detail_response = requests.get(aem_activity_url)
                print('{0} sheet, row_index is {1}, aem_activity_url is: {2}'.format(sheet_name, row_index,
                                                                                     aem_activity_url))
                assert_that(aem_activity_detail_response.status_code == 200)
                aem_activity_question_list = aem_activity_detail_response.json()['Questions']

                intend_recommendation_question = None
                if skill_type not in ('listening', 'reading'):
                    for question in aem_activity_question_list:
                        if question['key'].startswith(intend_recommendation):
                            intend_recommendation_question = question
                else:
                    intend_recommendation_question = aem_activity_question_list[0]

                if intend_recommendation_question is not None:
                    aem_question_tags = intend_recommendation_question['body']['tags']
                    if len(aem_question_tags) > 0:
                        intend_recommendation_tags = remove_non_used_tag_fields(
                            aem_question_tags[0])
                        sheet.cell(row=row_index, column=12).value = json.dumps(intend_recommendation_tags)

                        mistaken_question_tags = json.loads(row[4].value)
                        key_diff_list = get_tag_dict_difference(mistaken_question_tags,
                                                                     intend_recommendation_tags)
                        sheet.cell(row=row_index, column=13).value = "\r\n".join(key_diff_list)

    my_wb.save(file_name)


if __name__ == '__main__':
    source_aem_env = 'LIVE'
    release_program = 'Highflyers35'
    check_book_list = ["book-1", "book-2", "book-3", "book-4", "book-5", "book-6", "book-7", "book-8"]
    region_ach = "cn-3"
    get_mistaken_question_remediations(source_aem_env, release_program, check_book_list, region_ach)
    print("-----end of the check-----")
